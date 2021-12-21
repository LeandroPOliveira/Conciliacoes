from tkinter import *
from tkinter import ttk
import tkinter.messagebox
from tkinter import filedialog as fd
import pandas as pd
from datetime import datetime
from dateutil.relativedelta import relativedelta
import openpyxl
import os
import numpy as np
from reportlab.pdfgen import canvas
from PyPDF2 import PdfFileWriter, PdfFileReader
from win32com import client
import threading


class Conciliacao:

    def __init__(self, janela):
        self.janela = janela
        self.janela.title('Conciliações')
        self.janela.geometry('1000x600+200+50')

        self.lista_usuarios = []
        with open('usuarios.txt') as user:
            usuarios = user.readlines()
            for u in usuarios:
                self.lista_usuarios.append(u.strip())


        self.frame1 = Frame(self.janela, width=800, height=400, bg='white', bd=5, relief=RIDGE).\
            grid(padx=100, pady=100)
        Label(self.frame1, text='Login', font=('Impact', 35, 'bold'), fg='#6162FF', bg='white').\
            place(x=440, y=130)
        Label(self.frame1, text='Usuario', font=('Goudy old style', 15, 'bold'), fg='grey', bg='white'). \
            place(x=400, y=220)

        self.fonte = ('Goudy old style', 14)
        self.janela.option_add('*TCombobox*Listbox.font', self.fonte)
        self.usuario = ttk.Combobox(self.frame1, font=('arial', 14, 'bold'), width=17)
        self.usuario['values'] = ([i for i in self.lista_usuarios])
        # self.usuario.current(0)
        self.usuario.place(x=400, y=250)

        self.btn_entrar = Button(self.frame1, text='Entrar', font=('Goudy old style', 15, 'bold'), width=10,
                                 bd=0, bg='#6162FF', fg='white', command=self.tela_inicial)
        self.btn_entrar.place(x=440, y=350)

    def tela_inicial(self):
        # self.janela.withdraw()
        self.inicio = Toplevel()
        self.inicio.geometry('1000x600+200+50')

        self.tela_frame = Frame(self.inicio, width=800, height=500, bg='white')
        self.tela_frame.place(x=100, y=50)
        Label(self.tela_frame, text='Selecione a Competência', fg='grey', bg='white', font=('Goudy old style', 15, 'bold')). \
            place(x=280, y=50)
        lista = []
        for i in range(12):
            mes = datetime.today()
            data_limite = mes - relativedelta(months=i)
            lista.append(data_limite.strftime('%m.%Y'))


        self.competencia = ttk.Combobox(self.tela_frame, font=('arial', 16, 'bold'), width=15)
        self.competencia['values'] = (lista)
        self.inicio.option_add('*TCombobox*Listbox.font', self.fonte)
        # self.competencia.current(0)
        self.competencia.place(x=300, y=100)
        self.status = Button(self.tela_frame, width=15, text='Verificar', font=('Goudy old style', 15, 'bold'),
                               bd=1, bg='#6162FF', fg='white', command=self.status)
        self.status.place(x=200, y=160)

        self.progress_frame = Frame(self.inicio, width=400, height=200, bg='azure', bd=3, relief=RIDGE)
        self.texto = Label(self.progress_frame, text='Carregando...', font=('Goudy old style', 12, 'bold'), bd=0).\
            place(x=140, y=60)
        self.pb = ttk.Progressbar(self.progress_frame, mode='indeterminate', length=280)
        self.pb.place(x=50, y=100)

        def start_foo_thread(processo):
            # global foo_thread
            self.foo_thread = threading.Thread(target=processo)
            self.foo_thread.daemon = True
            self.progress_frame.place(x=300, y=200)
            self.pb.start()
            self.foo_thread.start()
            self.inicio.after(20, check_foo_thread)

        def check_foo_thread():
            if self.foo_thread.is_alive():
                self.inicio.after(20, check_foo_thread)
            else:
                self.pb.stop()
                self.progress_frame.place_forget()



        self.verifica = Button(self.tela_frame, width=15, text='Gerar Relatório', font=('Goudy old style', 15, 'bold'),
                               bd=1, bg='#6162FF', fg='white', command=lambda: start_foo_thread(self.validacao))
        self.verifica.place(x=400, y=160)

        self.nome1 = Label(self.tela_frame, text=self.lista_usuarios[0].strip(), font=('Goudy old style', 12), relief='groove',
                             width=20, height=1, bg='white', anchor=W).place(x=200, y=270)
        self.nome2 = Label(self.tela_frame, text=self.lista_usuarios[1], font=('Goudy old style', 12), relief='groove',
                             width=20, height=1, bg='white', anchor=W).place(x=200, y=295)
        self.nome3 = Label(self.tela_frame, text=self.lista_usuarios[2], font=('Goudy old style', 12), relief='groove',
                             width=20, height=1, bg='white', anchor=W).place(x=200, y=320)

        self.my_canvas1 = tkinter.Canvas(self.tela_frame, width=17, height=17, bg='white')  # Create 200x200 Canvas widget
        self.my_canvas1.place(x=390, y=272)
        self.my_oval1 = self.my_canvas1.create_oval(2, 2, 16, 16)

        self.my_canvas2 = tkinter.Canvas(self.tela_frame, width=17, height=17, bg='white', relief='groove')  # Create 200x200 Canvas widget
        self.my_canvas2.place(x=390, y=297)
        self.my_oval2 = self.my_canvas2.create_oval(2, 2, 16, 16)
        # self.my_canvas2.itemconfig(self.my_oval2, fill="red")
        self.my_canvas3 = tkinter.Canvas(self.tela_frame, width=17, height=17, bg='white',
                                    relief='groove')  # Create 200x200 Canvas widget
        self.my_canvas3.place(x=390, y=322)
        self.my_oval3 = self.my_canvas3.create_oval(2, 2, 16, 16)
        # my_canvas3.itemconfig(my_oval3, fill="green")

        self.status1 = Label(self.tela_frame, text='', font=('Goudy old style', 12), relief='groove',
                             width=20, height=1, bg='white')
        self.status1.place(x=410, y=270)
        self.status2 = Label(self.tela_frame, text='', font=('Goudy old style', 12), relief='groove',
                             width=20, height=1, bg='white')
        self.status2.place(x=410, y=295)
        self.status3 = Label(self.tela_frame, text='', font=('Goudy old style', 12), relief='groove',
                             width=20, height=1, bg='white')
        self.status3.place(x=410, y=320)





        def assina_gestor():
            with open('dados.txt', 'r') as f:
                lines = f.readlines()
                valid_user = []
                for i in lines:
                    i = i.split(';')
                    if i[0] == self.competencia.get() and i[1] == self.lista_usuarios[0] and i[2].strip() == 'OK':
                        valid_user.append('OK')
                    if i[0] == self.competencia.get() and i[1] == self.lista_usuarios[1] and i[2].strip() == 'OK':
                        valid_user.append('OK')
                    if i[0] == self.competencia.get() and i[1] == self.lista_usuarios[2] and i[2].strip() == 'OK':
                        valid_user.append('OK')

            if len(valid_user) >= 3:
                self.caminho = 'G:\GECOT\CONCILIAÇÕES CONTÁBEIS\CONCILIAÇÕES_' + self.competencia.get()[3:7] + \
                               '\\' + self.competencia.get()
                c = canvas.Canvas('watermark.pdf')
                # Draw the image at x, y. I positioned the x,y to be where i like here
                c.drawImage(self.usuario.get() + '.png', 350, 40, 150, 90,
                            mask='auto')
                c.save()
                watermark = PdfFileReader(
                    open("C:\\Users\loliveira\PycharmProjects\Excel\\watermark.pdf", "rb"))

                lista = []
                # arquivo_zip = ZipFile(self.caminho + '\\' + self.competencia.get() + '.zip', 'a')
                for file in os.listdir(self.caminho):
                    if file.endswith(".pdf"):
                        output_file = PdfFileWriter()
                        with open(self.caminho + '\\' + file, "rb") as f:
                            input_file = PdfFileReader(f, "rb")
                            # Number of pages in input document
                            page_count = input_file.getNumPages()

                            # Go through all the input file pages to add a watermark to them
                            for page_number in range(page_count):
                                input_page = input_file.getPage(page_number)
                                if page_number == page_count - 1:
                                    input_page.mergePage(watermark.getPage(0))
                                output_file.addPage(input_page)

                            with open(self.caminho + '\\' + file[8:], "wb") as outputStream:
                                output_file.write(outputStream)

                        os.remove(self.caminho + '\\' + file)
                        # arquivo_zip.write(self.caminho + '\\' + file)

                # arquivo_zip.close()
                tkinter.messagebox.showinfo('', 'Conciliações Assinadas com Sucesso!')

            else:
                tkinter.messagebox.showwarning('', 'Validação Pendente')


        if self.usuario.get() == 'Paulo França':
            self.gestor = Button(self.tela_frame, text='Assinar', font=('Goudy old style', 15, 'bold'),
            width=10, bd=1, bg='#6162FF', fg='white', command=assina_gestor).place(x=330, y=380)



    def status(self):
        with open('dados.txt', 'r') as f:
            lines = f.readlines()
            self.status1.config(text='')
            self.status2.config(text='')
            self.status3.config(text='')
            for i in lines:
                i = i.split(';')
                if i[0] == self.competencia.get() and i[1] == self.lista_usuarios[0] and i[2].strip() == 'OK':
                    self.status1.config(text='Validado', bg='light green')
                    self.my_canvas1.itemconfig(self.my_oval1, fill="green")
                elif i[0] == self.competencia.get() and i[1] == self.lista_usuarios[1] and i[2].strip() == 'OK':
                    self.status2.config(text='Validado', bg='light green')
                    self.my_canvas2.itemconfig(self.my_oval2, fill="green")
                elif i[0] == self.competencia.get() and i[1] == self.lista_usuarios[2] and i[2].strip() == 'OK':
                    self.status3.config(text='Validado', bg='light green')
                    self.my_canvas3.itemconfig(self.my_oval3, fill="green")
                else:
                    if self.status1.cget('text') == '':
                        self.status1.config(text='Validação Pendente', bg='light coral')
                        self.my_canvas1.itemconfig(self.my_oval1, fill="red")
                    if self.status2.cget('text') == '':
                        self.status2.config(text='Validação Pendente', bg='light coral')
                        self.my_canvas2.itemconfig(self.my_oval2, fill="red")
                    if self.status3.cget('text') == '':
                        self.status3.config(text='Validação Pendente', bg='light coral')
                        self.my_canvas3.itemconfig(self.my_oval3, fill="red")


    def validacao(self):
        self.caminho = 'G:\GECOT\CONCILIAÇÕES CONTÁBEIS\CONCILIAÇÕES_' + self.competencia.get()[3:7] + \
                       '\\' + self.competencia.get()
        pasta1 = os.listdir(self.caminho)

        lista = [[], [], [], [], []]

        for i in pasta1[::-1]:
            if i.startswith('~') == True:
                pasta1.remove(i)


        for i in pasta1:
            if i.endswith('.xlsx'):
                wb = openpyxl.load_workbook(self.caminho + '\\' + i, read_only=True)
                sheets = wb.sheetnames
                ws = wb[sheets[0]]
                try:
                    conta = ws['A2'].value.split()
                except:
                    conta = ['', '']
                valor_deb = ws['C5'].value
                valor_cred = ws['D5'].value
                try:
                    data1 = ws['A5'].value.strftime('%m.%Y')
                except:
                    data1 = ws['A5'].value
                lista[0].append(conta[1])
                lista[1].append(data1)
                lista[2].append(valor_deb)
                lista[3].append(valor_cred)

                wb.close()

        data = pd.DataFrame(lista).T

        data.columns = ['Conta', 'Data', 'Debito', 'Credito', 'Balancete']

        # Listar planilhas dos balancetes
        try:
            self.pasta = 'G:\GECOT\CONCILIAÇÕES CONTÁBEIS\CONCILIAÇÕES_' + self.competencia.get()[3:7] + \
                    '\BALANCETES\SOCIETÁRIOS\\'

            lista = os.listdir(self.pasta)

            lista4 = {}
            for i in lista:
                if self.competencia.get() in i or self.competencia.get().replace('.', '') in i:
                    tempo = os.path.getmtime(self.pasta + i)
                    tempo2 = datetime.fromtimestamp(tempo)
                    lista4.update({i: tempo2})

            dados = list(lista4.keys())[list(lista4.values()).index(max(lista4.values()))]
            dados = pd.read_excel(self.pasta + dados, skiprows=12)

        except:
            tkinter.messagebox.showinfo('', 'Balancete não encontrado.\n Informar local do arquivo.')
            dados = fd.askopenfilename(title='Abrir arquivo', initialdir=self.pasta)
            dados = pd.read_excel(dados, skiprows=12)



        dados = pd.DataFrame(dados)

        apoio = pd.read_excel('contas.xlsx')
        apoio = pd.DataFrame(apoio)

        for index1, row1 in data.iterrows():
            for index, row in dados.iterrows():
                if row1['Conta'] == row['Conta CSPE']:
                    # data.insert(3, 'Saldo', '')
                    data['Balancete'].loc[index1] = dados.loc[index, ' Saldo Acumulado']

        data[['Debito', 'Credito', 'Balancete']] = data[['Debito', 'Credito', 'Balancete']].apply(pd.to_numeric)

        data.fillna(0, inplace=True)
        data = data.round(2)
        data['Conciliação'] = data['Debito'] - data['Credito']

        data.drop(['Debito', 'Credito'], axis=1, inplace=True)

        data['Diferença'] = data['Conciliação'] - data['Balancete']

        # data['Resultado'] = data.apply(lambda x: x['Debito'] - x['Saldo'], axis=1)
        data = pd.merge(data, apoio[['Conta', 'Usuario']], on=['Conta'], how='left')

        data['Status'] = np.where(data['Diferença'] != 0, 'Diferença de Valor',
                                  (np.where(data['Data'] != self.competencia.get(), 'Data Incorreta', 'OK')))

        if self.usuario.get() != self.lista_usuarios[3]:
            data = data.loc[data['Usuario'] == self.usuario.get()]


        # data.to_excel('teste.xlsx', index=False)

        self.data = data

        try:
            self.relat.destroy()
            self.relatorio()
        except:
            self.relatorio()



    def relatorio(self):
        self.inicio.withdraw()
        self.relat = Toplevel()
        self.relat.geometry('1000x600+200+50')

        self.val_frame = Frame(self.relat, width=800, height=400, bg='white')
        self.val_frame.place(x=100, y=100)

        estilo = ttk.Style()
        estilo.theme_use('default')
        estilo.configure('Treeview', background='#D3D3D3', foreground='black', rowheight=25,
                         fieldbackground='#D3D3D3')
        estilo.map('Treeview', background=[('selected', '#347083')])

        # Treeview frame
        tree_frame = Frame(self.val_frame)
        tree_frame.pack(pady=50)
        # Barra rolagem
        tree_scroll = Scrollbar(tree_frame)
        tree_scroll.pack(side=RIGHT, fill=Y)
        # Criar Treeview
        nf_tree = ttk.Treeview(tree_frame, yscrollcommand=tree_scroll.set, selectmode='extended')
        nf_tree.pack(side=LEFT)
        # Configurar Barra Rolagem
        tree_scroll.config(command=nf_tree.yview)
        # Definir colunas
        colunas2 = ['Conta', 'Data', 'Balancete', 'Conciliação', 'Diferença', 'Usuario', 'Status']
        nf_tree['columns'] = colunas2
        # formatar colunas
        nf_tree.column('Conta', width=100)
        nf_tree.column('Data', width=100)
        nf_tree.column('Balancete', width=100)
        nf_tree.column('Conciliação', width=100)
        nf_tree.column('Diferença', width=100)
        nf_tree.column('Usuario', width=140)
        nf_tree.column('Status', width=140)

        # formatar títulos
        nf_tree.heading('Conta', text='Conta', anchor=W)
        nf_tree.heading('Data', text='Data', anchor=W)
        nf_tree.heading('Balancete', text='Balancete', anchor=W)
        nf_tree.heading('Conciliação', text='Conciliação', anchor=W)
        nf_tree.heading('Diferença', text='Diferença', anchor=W)
        nf_tree.heading('Usuario', text='Usuario', anchor=W)
        nf_tree.heading('Status', text='Status', anchor=W)

        nf_tree['show'] = 'headings'

        self.btn_validar = Button(self.relat, text='Revalidar', font=16, command=self.validacao)
        self.btn_validar.place(x=350, y=500)
        self.btn_assinar = Button(self.relat, text='Assinar', font=16, command=self.assinar)
        self.btn_assinar.place(x=550, y=500)

        photo = PhotoImage(file='botao.png')
        # photoimage = photo.subsample(3, 3)
        self.btn_voltar = Button(self.relat, text='Voltar', font=('arial', 16), bd=2,
                                 command=lambda: [self.relat.withdraw(), self.inicio.deiconify()])
        self.btn_voltar.place(x=20, y=20)


        # inserir dados do banco no treeview
        def inserir_tree(lista):
            nf_tree.delete(*nf_tree.get_children())
            contagem = 0
            for index, row in lista.iterrows():  # loop para inserir cores diferentes nas linhas
                if row[6] == 'OK':
                    nf_tree.insert(parent='', index='end', text='', iid=contagem,
                                   values=(row[0], row[1], row[2], row[3], row[4], row[5], row[6]), tags=('evenrow',))
                else:
                    nf_tree.insert(parent='', index='end', text='', iid=contagem,
                                   values=(row[0], row[1], row[2], row[3], row[4], row[5], row[6]), tags=('oddrow',))
                contagem += 1

        nf_tree.tag_configure('oddrow', background='light coral')
        nf_tree.tag_configure('evenrow', background='lightgreen')
        inserir_tree(self.data)

        def NotasInfo2(ev):
            verinfo2 = nf_tree.focus()
            dados2 = nf_tree.item(verinfo2)
            row = dados2['values']
            os.startfile(self.caminho + '\\' + 'Conta ' + row[0].replace('.', '') + '.xlsx')

        nf_tree.bind('<Double-Button>', NotasInfo2)

    def assinar(self):
        valida = self.data['Status'].unique()

        if 'OK' in valida and len(valida) == 1:
            lista3 = []

            for i in self.data['Conta']:
                conta = 'Conta ' + i.replace('.', '') + '.xlsx'
                lista3.append(conta)
                print(lista3)

            # Create the watermark from an image
            c = canvas.Canvas('watermark.pdf')
            # Draw the image at x, y. I positioned the x,y to be where i like here
            c.drawImage(self.usuario.get() + '.png', 40, 50, 150, 100,
                        mask='auto')
            c.save()

            for i in lista3:

                # Open Microsoft Excel
                excel = client.Dispatch("Excel.Application")

                # Read Excel File
                sheets = excel.Workbooks.Open(self.caminho + '\\' + i)
                work_sheets = sheets.Worksheets[0]

                # Convert into PDF File
                path = 'C:\\Users\loliveira\PycharmProjects\Excel\\' + 'teste ' + i.replace('.xlsx', '.pdf')

                work_sheets.ExportAsFixedFormat(0, path)


                # Get the watermark file you just created
                watermark = PdfFileReader(open("watermark.pdf", "rb"))
                # Get our files ready

                output = PdfFileWriter()

                with open(path, "rb") as provisorio:
                    input = PdfFileReader(provisorio)
                    number_of_pages = input.getNumPages()


                    for current_page_number in range(number_of_pages):
                        page = input.getPage(current_page_number)
                        if page.extractText() != "":
                            output.addPage(page)

                    page_count = output.getNumPages()
                    output2 = PdfFileWriter()
                    # Go through all the input file pages to add a watermark to them
                    for page_number in range(page_count):
                        input_page = output.getPage(page_number)
                        if page_number == page_count - 1:
                            input_page.mergePage(watermark.getPage(0))
                        output2.addPage(input_page)

                    # finally, write "output" to document-output.pdf
                    with open(self.caminho + '\\' + 'pendente' + i.replace('.xlsx', '.pdf'), "wb") as outputStream:
                        output2.write(outputStream)


                os.remove(path)
                sheets.Close(True)
                # excel.Quit()


            adicionar = [self.competencia.get(), self.usuario.get(), 'OK']
            adicionar = ';'.join(adicionar)

            with open('dados.txt', 'a') as f:
                f.write(f'\n{adicionar}')

            tkinter.messagebox.showinfo('', 'Arquivos assinados com Sucesso!')

        else:
            tkinter.messagebox.showwarning('', 'Erro! Verificar pendências!')



if __name__=='__main__':
    janela = Tk()
    aplicacao = Conciliacao(janela)
    janela.mainloop()
